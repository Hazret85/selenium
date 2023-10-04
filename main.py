from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
import openpyxl


workbook = openpyxl.Workbook()
sheet = workbook.active


def result_students():
    try:
        result = []
        schedule_elements = driver.find_elements(By.CLASS_NAME,'_body')
        time.sleep(3)
        for tablo in schedule_elements:
            info = [x for x in tablo.text.split('\n') if len(x)>2]
            sheet.append(info)
            result.append(info)
        percentages = [int(item[1].rstrip('%')) for item in result if len(item) == 2]
        average_percentage = sum(percentages) / len(percentages) if percentages else 0
        sheet.append(['Средняя успеваемость группы:',average_percentage])
    except Exception as error:
        print(error)

        # Сохраняем файл

name = teacher_name
date_1 = '2023-09-2'
date_2 = '202023-09-3'
service = Service(executable_path='chromedriver.exe')
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)
driver.get('https://lms.algoritmika.org/group/default/schedule?GroupLessonSearch%5Bstart_time%5D={}%20-%{}&GroupLessonSearch%5Bnumber%5D=&GroupLessonSearch%5Blesson.title%5D=&GroupLessonSearch%5Bgroup_id%5D=&GroupLessonSearch%5Bgroup.title%5D=&GroupLessonSearch%5Bgroup.venue%5D=&GroupLessonSearch%5Bgroup.active_student_count%5D=&GroupLessonSearch%5Bweekday%5D=&GroupLessonSearch%5Bteacher.name%5D={}&GroupLessonSearch%5Bis_online%5D='.format(date_1,date_2,name))
time.sleep(3)
login_input= driver.find_element(By.ID, 'login')
login_input.send_keys(log)
login_input.send_keys(Keys.ENTER)
time.sleep(2)
login_input= driver.find_element(By.ID, 'password')
login_input.send_keys(pas)
login_input.send_keys(Keys.ENTER)
time.sleep(5)
elements = driver.find_elements(By.CSS_SELECTOR,'td.schedule-grid[data-col-seq="group_id"]')
title = driver.find_elements(By.CSS_SELECTOR,'td.schedule-grid[data-col-seq="lesson.title"]')

data = {}
for x in range(len(elements)):
    if title[x].text.find('МК') == -1:
        data[elements[x].text] = title[x].text


for id_g in data.keys():
    driver.get('https://lms.algoritmika.org/group/view/{}'.format(id_g))
    time.sleep(3)

    link = driver.find_element(By.LINK_TEXT,'Успехи')
    link.click()
    time.sleep(3)

    filename = 'group_id_{}.xlsx'.format(id_g)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    headers = ['Имя', 'Прогресс']


    url = driver.current_url
    sheet.append([url])
    sheet.append([data[id_g]])
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=3, column=col, value=header)
    for x in range(4):
        result_students()
        current_url = driver.current_url
        first = current_url.find('Lesson')
        less_number = current_url[first + 9:70]
        driver.get(current_url.replace(less_number,str(int(less_number)-1)))
        time.sleep(3)
        sheet.append(['_______________________'])
        sheet.append([driver.current_url])
    workbook.save(filename)

driver.quit()
